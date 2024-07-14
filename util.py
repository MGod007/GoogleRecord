import pandas as pd
import  openpyxl
from openpyxl import Workbook

def readexcel():
    path = "SearchKeywords.xlsx"
    excel_obj = openpyxl.load_workbook(path)
    sheet_obj = excel_obj.active
    row = sheet_obj.max_row
    column = sheet_obj.max_column
    listofdata = []
    for i   in range(2, row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        listofdata.append(cell_obj.value)
        
    return   listofdata

def write_to_excel(data):
    path = "GoogleSearchResult.xlsx"
    excel_obj = openpyxl.load_workbook(path)
    sheet_obj = excel_obj.active
    
    result = list(data)
    count = len(result)
    for i  in range(1, count):
        sheet_obj.cell(row=1, column=1).value = "Records"
        sheet_obj.cell(row=i+1, column=1).value = result[i]
    excel_obj.save(path)


    


    
